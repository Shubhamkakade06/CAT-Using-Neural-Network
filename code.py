import numpy as np
import pandas as pd
import torch
import torch.nn.parallel
import torch.utils.data
from openpyxl import Workbook
import math

dataset=pd.read_csv('AI-DataTrain.csv')
finaltest=pd.read_csv('AI-DataTest.csv')

train=dataset.iloc[:900,1:51]
test=dataset.iloc[900:,1:51]

nb_users=900
nb_ques=50

training_set=train.values.tolist()

test_set=test.values.tolist()

training_set= torch.FloatTensor(training_set)
test_set= torch.FloatTensor(test_set)

class RBM():
  def __init__(self,nv,nh):
    self.W = torch.randn(nh,nv)
    self.a = torch.randn(1,nh)
    self.b = torch.randn(1,nv)

  def sample_h(self,x):
    wx = torch.mm(x, self.W.t())
    activation = wx + self.a.expand_as(wx)
    p_h_given_v = torch.sigmoid(activation)
    return p_h_given_v, torch.bernoulli(p_h_given_v)

  def sample_v(self,y):
    wy = torch.mm(y, self.W)
    activation = wy + self.b.expand_as(wy)
    p_v_given_h = torch.sigmoid(activation)
    return p_v_given_h, torch.bernoulli(p_v_given_h)

  def train(self,v0,vk,ph0,phk):
    self.W += torch.mm(v0.t(), ph0) - torch.mm(vk.t(),phk)
    self.b += torch.sum((v0 - vk),0)
    self.a += torch.sum((ph0 - phk),0)


nv = len(training_set[0])
nh = 50
batch_size = 75
rbm = RBM(nv,nh)


#training the RBM
nb_epoch=10
for epoch in range(1 , nb_epoch + 1):
  tain_loss = 0
  s = 0.
  for id_user in range(0, nb_users - batch_size, batch_size):
    vk = training_set[id_user:id_user+batch_size]
    v0 = training_set[id_user:id_user+batch_size]
    ph0,_ = rbm.sample_h(v0)
    for k in range(10):
      _,hk = rbm.sample_h(vk)
      _,vk = rbm.sample_v(hk)
      vk[v0<0] = v0[v0<0]
    phk,_ = rbm.sample_h(vk)
    rbm.train(v0, vk, ph0, phk)
    train_loss += torch.mean(torch.abs(v0[v0>=0] - vk[v0>=0]))
    s += 1.

#testing the RBM
test_loss = 0
s = 0.
for id_user in range(nb_users):
  v = training_set[id_user:id_user+1]
  vt = test_set[id_user:id_user+1]
  if len(vt[vt>=0]) > 0:
    _,h = rbm.sample_h(v)
    _,v = rbm.sample_v(h)
    test_loss += torch.mean(torch.abs(vt[vt>=0] - v[vt>=0]))
    s += 1.
#print('test_loss: ' +str(test_loss/s))

result=rbm.W

result=result.tolist()

final_result=[]
for i in result:
    k=0
    add=0
    for j in i:
        add=j+add
        k+=1
    avg=add/k
    final_result.append(avg)

mylist=[]
for i in final_result:
  if i<0:
    mylist.append(float("-"+str(math.sqrt(math.sqrt(abs(i))))))
  else:
    mylist.append(math.sqrt(math.sqrt(i)))

minimum=min(mylist)
scale=abs(minimum)+1

newlist=[]
for i in mylist:
  newlist.append(i+scale)

wb = Workbook()

ws =  wb.active
ws.title = "Final Weights"


c1 = ws.cell(row = 1, column = 1 )
c1.value = ('Quetions')
c2 = ws.cell(row = 1, column = 2 )
c2.value =('Weightage')
j=2
for i in newlist:
    c1 = ws.cell(row = j, column = 1 )
    c1.value = ('Q'+str(j))
    c2 = ws.cell(row = j, column = 2 )
    c2.value = newlist[j-2]
    j += 1

wb.save(filename = 'output.xlsx')